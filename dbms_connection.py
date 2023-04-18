import psycopg2
import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active
# set the column headers and format them
headers = [
    "matcode",
    "title",
   "ref",
    "au",
    "safstk",
    "reorder",
   "ar",
    "desk",
    "ordcst",
    "eoq",
    "lt",
    "safty",
    "auamt",
    "spare",
    "gr",
   "nm",
   "pstock",
   "ind",
   "nsaftystk",
   "specno",
   "matgroup",
   "section",
   "group_b",
    "group_c",
    "group_a",
   "abc",
   "reordqty",
    "unitrate",
    "dwgno",
]
for col, header in enumerate(headers, start=1):
    cell = worksheet.cell(row=1, column=col)
    cell.value = header
    cell.font = openpyxl.styles.Font(bold=True)
    cell.alignment = openpyxl.styles.Alignment(horizontal="center")





#establishing the connection
conn = psycopg2.connect(
   database="bheleml", user='tester', password='Password', host='127.0.0.1', port= '5432'
)

#Setting auto commit false
conn.autocommit = True

#Creating a cursor object using the cursor() method
cursor = conn.cursor()

#Retrieving data
cursor.execute('''SELECT * from bhelksgd.mtlsmast''')

#Fetching 1st row from the table

#Fetching 1st row from the table
data = cursor.fetchall();
print(data)

conn.commit()
# populate the rows in the Excel file
for row, record in enumerate(data, start=2):
    for col, value in enumerate(record, start=1):
        cell = worksheet.cell(row=row, column=col)
        cell.value = value
# save the Excel file
workbook.save("material.xlsx")

#Closing the connection
conn.close()